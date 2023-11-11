import re
from PyPDF2 import PdfReader
import openpyxl
import pandas as pd

# Criar listas para armazenar os resultados
dates_boleto = []
valor_boleto = []
beneficiario_boleto = []
codigo_boleto = []
numero_boleto = []

# Abra o arquivo PDF
with open('C:\oletos\este.pdf', 'rb') as pdf_file:
    pdf_reader_bb = PdfReader(pdf_file)

    # Itere pelas páginas do PDF
    for page_num, page in enumerate(pdf_reader_bb.pages):
        page_text = page.extract_text()

        # Padrão para data
        date_pattern = r'\d{2}/\d{2}/\d{4}'
        dates = re.findall(date_pattern, page_text)

        # Padrão para valores monetários
        value_pattern = r'\(=\) Valor Cobrado\s*([\d.,]+)'
        values = re.findall(value_pattern, page_text)

        # Padrão para beneficiário
        beneficiario_pattern = r'Beneficiário\s*([^Espécie]+)'
        beneficiario = re.findall(beneficiario_pattern, page_text)

        # Padrão para o código de barras
        codigo_de_barras_pattern = r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d{1} \d{14})'
        codigo_de_barras = re.findall(codigo_de_barras_pattern, page_text)

        numero_documento_pattern = r'(\d{4}-\d{1}/\d{1})'
        numero_documento = re.findall(numero_documento_pattern, page_text)

        if dates:
            dates_boleto.append(f'{dates[0]}')

        if values:
            valor_boleto.append(f'{values[0]}')

        if beneficiario:
            beneficiario_text = beneficiario[0].strip()
            beneficiario_boleto.append(f'{beneficiario_text}')

        if codigo_de_barras:
            codigo_boleto.append(f'{codigo_de_barras[0]}')

        if numero_documento:
            numero_boleto.append(f' {numero_documento[0]}')

# Crie um DataFrame a partir das listas
data = {
    '': '',
    'Empresa': beneficiario_boleto,
    'Código de Barras': codigo_boleto,
    'Vencimento': dates_boleto,
    'Valor': valor_boleto,
    'Mês': '',
    'NF': numero_boleto
}

df = pd.DataFrame(data)

# Carregar o arquivo Excel existente
excel_file = 'pedro.xlsm'
wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)
ws = wb.active  # Ou selecione uma planilha específica, se necessário

# Criar um conjunto (set) para armazenar os códigos de boleto existentes na planilha
existing_codes = set()

# Preencher o conjunto com os códigos de boleto existentes na planilha
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        existing_codes.add(cell.value)

# Converter o DataFrame em uma lista de listas
data_to_append = df.values.tolist()

# Adicionar os dados ao arquivo Excel com verificação de códigos de boleto
for row_data in data_to_append:
    # O código de barras está na terceira coluna (índice 2)
    codigo_boleto = row_data[2]
    if codigo_boleto not in existing_codes:
        ws.append(row_data)
        # Adicionar o código ao conjunto de códigos existentes
        existing_codes.add(codigo_boleto)

wb.save(excel_file)

wb.close()


dates_boleto = []
valor_boleto = []
beneficiario_boleto = []
codigo_boleto = []
numero_boleto = []

with open('C:\oletos\este2.pdf', 'rb') as pdf_file:
    pdf_reader_bb_correios = PdfReader(pdf_file)

    # Itere pelas páginas do PDF
    for page_num, page in enumerate(pdf_reader_bb_correios.pages):
        page_text = page.extract_text()

        date_pattern = r'\d{2}/\d{2}/\d{4}'
        dates = re.findall(date_pattern, page_text)

        value_pattern = r'R\$\s*([\d.,]+)'
        values = re.findall(value_pattern, page_text)

        beneficiario_pattern = r'Beneficiário\s*([^\n]+)'
        beneficiario = re.findall(beneficiario_pattern, page_text)

        codigo_de_barras_pattern = r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d{1} \d{14})'
        codigo_de_barras = re.findall(codigo_de_barras_pattern, page_text)

        numero_documento_pattern = r'(\d{6})'
        numero_documento = re.findall(numero_documento_pattern, page_text)

        # Verifique se encontrou alguma data e valor
        if len(dates) >= 2:
            segunda_data = dates[1]
            dates_boleto.append(f'{segunda_data}')

        if values:
            valor_boleto.append(f'{values[0]}')

        if beneficiario:
            beneficiario_boleto.append(f'{beneficiario[0]}')

        if codigo_de_barras:
            codigo_boleto.append(f'{codigo_de_barras[0]}')

        if numero_documento:
            numero_boleto.append(f'{numero_documento[0]}')
data = {
    '': '',
    'Empresa': beneficiario_boleto,
    'Código de Barras': codigo_boleto,
    'Vencimento': dates_boleto,
    'Valor': valor_boleto,
    'Mês': '',
    'NF': numero_boleto
}

df = pd.DataFrame(data)


# Carregar o arquivo Excel existente
excel_file = 'pedro.xlsm'
wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)
ws = wb.active  

# Criar um conjunto (set) para armazenar os códigos de boleto existentes no Excel
existing_codes = set()

# Preencher o conjunto com os códigos de boleto existentes no Excel
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        existing_codes.add(cell.value)


for index, row in df.iterrows():
    codigo_boleto = row['Código de Barras']
    if codigo_boleto not in existing_codes:
        ws.append(list(row))
        existing_codes.add(codigo_boleto)


wb.save(excel_file)


wb.close()


dates_boleto = []
valor_boleto = []
beneficiario_boleto = []
codigo_boleto = []
numero_boleto = []


with open('C:\oletos\este3.pdf', 'rb') as pdf_file:
    pdf_reader_san = PdfReader(pdf_file)

    # Itere pelas páginas do PDF
    for page_num, page in enumerate(pdf_reader_san.pages):
        page_text = page.extract_text()


        date_pattern = r'\d{2}/\d{2}/\d{4}'
        dates = re.findall(date_pattern, page_text)

 
        value_pattern = r'Valor da fatura\s*R\$\s*([\d.,]+)'
        values = re.findall(value_pattern, page_text)

        beneficiario_pattern = r'Beneficiário\s*([^\n]+)'
        beneficiario = re.findall(beneficiario_pattern, page_text)

        codigo_de_barras_pattern = r'(\d{5}\.\d{5} \d{5}\.\d{6} \d{5}\.\d{6} \d{1} \d{14})'
        codigo_de_barras = re.findall(codigo_de_barras_pattern, page_text)

        numero_documento_pattern = r'\d{7}-\d{2}'
        numero_documento = re.findall(numero_documento_pattern, page_text)

        # Verifique se encontrou alguma data e valor
        if len(dates) >= 2:
            segunda_data = dates[1]
            dates_boleto.append(f'{segunda_data}')

        if values:
            valor_boleto.append(f'{values[0]}')

        if beneficiario:
            beneficiario_inicio = beneficiario[0].split('- ')
            beneficiario_inicio = beneficiario_inicio[1].split(' /')
            beneficiario_boleto.append(f'{beneficiario_inicio[0]}')

        if codigo_de_barras:
            codigo_boleto.append(f'{codigo_de_barras[0]}')

        if numero_documento:
            numero_boleto.append(f'{numero_documento[0]}')

data = {
    '': '',
    'Empresa': beneficiario_boleto,
    'Código de Barras': codigo_boleto,
    'Vencimento': dates_boleto,
    'Valor': valor_boleto,
    'Mês': '',
    'NF': numero_boleto
}

df = pd.DataFrame(data)


# Carregar o arquivo Excel existente
excel_file = 'pedro.xlsm'
wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)
ws = wb.active  
existing_codes = set()

# Preencher o conjunto com os códigos de boleto existentes no Excel
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        existing_codes.add(cell.value)

# Iterar sobre os dados do DataFrame e adicionar apenas os que não estão no Excel
for index, row in df.iterrows():
    codigo_boleto = row['Código de Barras']
    if codigo_boleto not in existing_codes:
        ws.append(list(row))
        existing_codes.add(codigo_boleto)


wb.save(excel_file)


wb.close()
